"""
Correcciones al archivo Matriz_Comunicaciones_AMPA_2026.xlsx:

1. Resumen General:
   - Limpiar prefijos ' de celdas de datos
   - % Avance → fórmula IFERROR(AVERAGEIF(Actividades[Proyecto],A?,Actividades[% Avance (auto)]),0)
   - N° Actividades → =COUNTIF(Actividades[Proyecto], A?)
   - N° Indicadores → =COUNTIF(Indicadores[Proyecto], A?)
   - Extender tabla para que abarque filas adicionales vacías (expansión dinámica)

2. Actividades por Proyecto:
   - % Avance: usar ID construido como A?&"_"&(ROW()-header_row) igualando formato de TAREAS
   - Tareas totales: COUNTIF por ID correcto
   - Tareas completas: COUNTIFS por ID + "Completada"
   - El ID en TAREAS debe coincidir: PRY-01_1 = Proyecto_Nfila
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import copy

VERDE_SELVA = "1A3A2A"
VERDE_HOJA  = "3D7A52"
VERDE_SUAVE = "E8F5E9"
OCRE_CLARO  = "F0D080"
BLANCO      = "FFFFFF"
CREMA       = "FAF6EE"
GRIS_TEXTO  = "4A4A4A"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style='thin', color="CCCCCC")
    return Border(top=s, bottom=s, left=s, right=s)

wb = load_workbook('Matriz_Comunicaciones_AMPA_2026_v2.xlsx')

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 1: RESUMEN GENERAL — fix hardcoded values + add formulas
# ──────────────────────────────────────────────────────────────────────────────
ws1 = wb['Resumen General']

# Find table bounds
tbl_pry = ws1.tables['Proyectos']
print(f"Proyectos table ref: {tbl_pry.ref}")  # A9:K21

# Header row is row 9, data starts row 10
HEADER_ROW = 9
DATA_START  = 10

# Count existing data rows
data_rows = []
for row in ws1.iter_rows(min_row=DATA_START, max_row=DATA_START+50, min_col=1, max_col=1):
    for cell in row:
        v = cell.value
        if v and str(v).strip().startswith("PRY"):
            data_rows.append(cell.row)

last_data_row = max(data_rows) if data_rows else DATA_START + 9
print(f"Data rows: {data_rows}, last: {last_data_row}")

# ── Clean the apostrophe prefix from all data cells ──────────────────────────
# openpyxl stores text-prefixed numbers as plain strings; the ' is stored as
# quotePrefix=True on the cell. We just need to ensure numeric columns hold numbers.
for row_num in range(DATA_START, last_data_row + 1):
    # Col G (7) = % Avance — skip, will replace with formula
    # Col H (8) = N° Actividades — skip, will replace
    # Col I (9) = N° Indicadores — skip, will replace
    pass  # we'll overwrite cols G,H,I with formulas below

# ── Insert formulas for cols G, H, I ─────────────────────────────────────────
# Col G = % Avance → average of % Avance column in Actividades for this project
# Col H = N° Actividades → count rows in Actividades where Proyecto = this code
# Col I = N° Indicadores → count rows in Indicadores where Proyecto = this code

for row_num in range(DATA_START, last_data_row + 1):
    a_ref = f"A{row_num}"

    # % Avance: average of all activities' % avance for this project
    # Since col M in Actividades stores 0..1 values (fractions), multiply by 100 for display
    ws1.cell(row=row_num, column=7).value = (
        f'=IFERROR(AVERAGEIF(\'Actividades por Proyecto\'!$A:$A,{a_ref},'
        f'\'Actividades por Proyecto\'!$M:$M)*100,0)'
    )
    ws1.cell(row=row_num, column=7).number_format = '0"%"'

    # N° Actividades: count matching rows in Actividades table
    ws1.cell(row=row_num, column=8).value = (
        f'=COUNTIF(\'Actividades por Proyecto\'!$A:$A,{a_ref})-1'
    )
    # Subtract 1 to exclude the header row match (header "Proyecto" won't match PRY-xx so it's 0 always)
    # Safer: use COUNTIFS with a non-empty check
    ws1.cell(row=row_num, column=8).value = (
        f'=COUNTIFS(\'Actividades por Proyecto\'!$A:$A,{a_ref},'
        f'\'Actividades por Proyecto\'!$C:$C,"<>"&"")'
    )

    # N° Indicadores
    ws1.cell(row=row_num, column=9).value = (
        f'=COUNTIFS(\'Indicadores Comunicación\'!$A:$A,{a_ref},'
        f'\'Indicadores Comunicación\'!$C:$C,"<>"&"")'
    )

    # Apply consistent styling
    for col in [7, 8, 9]:
        c = ws1.cell(row=row_num, column=col)
        bg = CREMA if row_num % 2 == 0 else BLANCO
        c.fill = fill(bg)
        c.font = Font(name="Arial", size=10, color=GRIS_TEXTO, bold=(col==7))
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

# ── Extend table to include 10 empty rows for future projects ─────────────────
EXTRA_ROWS = 10
new_last = last_data_row + EXTRA_ROWS

# First unmerge the old instruction box so we can write over those rows
ws1.unmerge_cells("A22:K25")
# Clear old instruction content
for row_num in range(22, 26):
    for col in range(1, 12):
        c = ws1.cell(row=row_num, column=col)
        c.value = None
        c.fill  = fill(BLANCO)

# Style blank expansion rows
for row_num in range(last_data_row + 1, new_last + 1):
    for col in range(1, 12):
        c = ws1.cell(row=row_num, column=col)
        c.fill = fill(CREMA if row_num % 2 == 0 else BLANCO)
        c.font  = Font(name="Arial", size=10, color=GRIS_TEXTO)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border()
    # Pre-populate formulas for G,H,I in blank rows too
    a_ref = f"A{row_num}"
    ws1.cell(row=row_num, column=7).value = (
        f'=IFERROR(AVERAGEIF(\'Actividades por Proyecto\'!$A:$A,{a_ref},'
        f'\'Actividades por Proyecto\'!$M:$M)*100,0)'
    )
    ws1.cell(row=row_num, column=7).number_format = '0"%"'
    ws1.cell(row=row_num, column=8).value = (
        f'=COUNTIFS(\'Actividades por Proyecto\'!$A:$A,{a_ref},'
        f'\'Actividades por Proyecto\'!$C:$C,"<>"&"")'
    )
    ws1.cell(row=row_num, column=9).value = (
        f'=COUNTIFS(\'Indicadores Comunicación\'!$A:$A,{a_ref},'
        f'\'Indicadores Comunicación\'!$C:$C,"<>"&"")'
    )

# Update table reference to cover extra rows
ws1.tables.pop('Proyectos')  # remove old
new_tbl = Table(
    displayName="Proyectos",
    ref=f"A{HEADER_ROW}:K{new_last}"
)
new_tbl.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False
)
ws1.add_table(new_tbl)
print(f"Updated Proyectos table to A{HEADER_ROW}:K{new_last}")

# Move instruction box down if needed
# Find and clear old instruction and rewrite below new table
for row_num in range(last_data_row + 1, last_data_row + 5):
    for col in range(1, 12):
        c = ws1.cell(row=row_num, column=col)
        if c.value and "📌" in str(c.value):
            c.value = None

inst_row = new_last + 2
ws1.merge_cells(f"A{inst_row}:K{inst_row+3}")
c = ws1.cell(row=inst_row, column=1,
    value="📌  CÓMO AGREGAR UN NUEVO PROYECTO: Escribe en la siguiente fila vacía de la tabla (ya están preparadas). "
          "% Avance, N° Actividades y N° Indicadores se calculan automáticamente desde las otras hojas. "
          "Luego agrega las actividades en 'Actividades por Proyecto' y los indicadores en 'Indicadores Comunicación'.")
c.fill = fill(OCRE_CLARO)
c.font = Font(name="Arial", italic=True, color=VERDE_SELVA, size=9)
c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
c.border = thin_border()
ws1.row_dimensions[inst_row].height = 55

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 2: ACTIVIDADES — fix % Avance formula
# ──────────────────────────────────────────────────────────────────────────────
ws2 = wb['Actividades por Proyecto']

# The TAREAS sheet uses ID format: PRY-XX_N where N is the row number within
# 'Actividades por Proyecto' data (row 6 = activity 1, row 7 = activity 2, etc.)
# So activity in row R has ID = A_R & "_" & (R - 5)   [header is row 5]

ACT_HEADER_ROW = 5
ACT_DATA_START = 6

# Find last activity row
act_rows = []
for row in ws2.iter_rows(min_row=ACT_DATA_START, max_row=ACT_DATA_START+100, min_col=1, max_col=1):
    for cell in row:
        v = cell.value
        if v and str(v).strip().startswith("PRY"):
            act_rows.append(cell.row)

last_act_row = max(act_rows) if act_rows else ACT_DATA_START + 14
print(f"Activity rows: {act_rows}, last: {last_act_row}")

# For each activity row, the correct ID is:  [proyecto_code]_[row_index]
# row_index = row_num - ACT_HEADER_ROW  (so row 6 → index 1, row 7 → 2, etc.)

for row_num in act_rows:
    row_idx = row_num - ACT_HEADER_ROW  # 1-based activity index
    a_ref   = f"A{row_num}"
    id_expr = f'{a_ref}&"_{row_idx}"'   # e.g. "PRY-01_1"

    # Col M (13): % Avance = completed tasks / total tasks for this activity ID
    # Formula: =IFERROR( COUNTIFS(TAREAS[ID Actividad], id, TAREAS[Estado Tarea],"Completada")
    #                   / COUNTIF(TAREAS[ID Actividad], id), "" )
    ws2.cell(row=row_num, column=13).value = (
        f'=IFERROR('
        f'COUNTIFS(TAREAS[ID Actividad],{id_expr},TAREAS[Estado Tarea],"Completada")'
        f'/COUNTIF(TAREAS[ID Actividad],{id_expr}),"")'
    )
    ws2.cell(row=row_num, column=13).number_format = '0%'

    # Col N (14): Total tareas
    ws2.cell(row=row_num, column=14).value = (
        f'=COUNTIF(TAREAS[ID Actividad],{id_expr})'
    )

    # Col O (15): Tareas completadas
    ws2.cell(row=row_num, column=15).value = (
        f'=COUNTIFS(TAREAS[ID Actividad],{id_expr},TAREAS[Estado Tarea],"Completada")'
    )

    # Style
    for col in [13, 14, 15]:
        c = ws2.cell(row=row_num, column=col)
        bg = CREMA if row_num % 2 == 0 else BLANCO
        c.fill  = fill(bg)
        c.font  = Font(name="Arial", size=10, color=GRIS_TEXTO,
                       bold=(col == 13))
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

# ── Extend Actividades table too ──────────────────────────────────────────────
ACT_EXTRA = 15
new_last_act = last_act_row + ACT_EXTRA

# Unmerge old instruction box
ws2.unmerge_cells("A22:P24")
for row_num in range(22, 25):
    for col in range(1, 17):
        c = ws2.cell(row=row_num, column=col)
        c.value = None
        c.fill  = fill(BLANCO)

# Style blank rows and add formulas
ACT_NCOLS = 16
for row_num in range(last_act_row + 1, new_last_act + 1):
    for col in range(1, ACT_NCOLS + 1):
        c = ws2.cell(row=row_num, column=col)
        c.fill   = fill(CREMA if row_num % 2 == 0 else BLANCO)
        c.font   = Font(name="Arial", size=10, color=GRIS_TEXTO)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border()
    # Add formulas even in blank rows (they'll show "" when A is empty)
    row_idx = row_num - ACT_HEADER_ROW
    a_ref   = f"A{row_num}"
    id_expr = f'{a_ref}&"_{row_idx}"'
    ws2.cell(row=row_num, column=13).value = (
        f'=IFERROR('
        f'COUNTIFS(TAREAS[ID Actividad],{id_expr},TAREAS[Estado Tarea],"Completada")'
        f'/COUNTIF(TAREAS[ID Actividad],{id_expr}),"")'
    )
    ws2.cell(row=row_num, column=13).number_format = '0%'
    ws2.cell(row=row_num, column=14).value = (
        f'=COUNTIF(TAREAS[ID Actividad],{id_expr})'
    )
    ws2.cell(row=row_num, column=15).value = (
        f'=COUNTIFS(TAREAS[ID Actividad],{id_expr},TAREAS[Estado Tarea],"Completada")'
    )

# Update table ref
ws2.tables.pop('Actividades')
new_act_tbl = Table(
    displayName="Actividades",
    ref=f"A{ACT_HEADER_ROW}:P{new_last_act}"
)
new_act_tbl.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium7", showRowStripes=True)
ws2.add_table(new_act_tbl)
print(f"Updated Actividades table to A{ACT_HEADER_ROW}:P{new_last_act}")

# ── Update instruction note in Actividades ────────────────────────────────────
for row_num in range(last_act_row + 1, last_act_row + 5):
    for col in range(1, 17):
        c = ws2.cell(row=row_num, column=col)
        if c.value and "📌" in str(c.value):
            c.value = None

inst_row2 = new_last_act + 2
ws2.merge_cells(f"A{inst_row2}:P{inst_row2+3}")
c = ws2.cell(row=inst_row2, column=1,
    value="📌  CÓMO AGREGAR ACTIVIDADES: Escribe en la siguiente fila vacía de la tabla. "
          "El ID de actividad para TAREAS es [Código Proyecto]_[N° de fila en esta hoja], ej: PRY-01_3 = fila 3 de datos. "
          "% Avance, Tareas totales y Completadas se calculan automáticamente desde la hoja TAREAS.")
c.fill  = fill(OCRE_CLARO)
c.font  = Font(name="Arial", italic=True, color=VERDE_SELVA, size=9)
c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
c.border = thin_border()
ws2.row_dimensions[inst_row2].height = 55

# ── Also extend TAREAS table to handle new tasks ─────────────────────────────
ws3 = wb['Tareas']
tbl_tsk = ws3.tables['TAREAS']
print(f"TAREAS table ref: {tbl_tsk.ref}")

# Find last TAREAS row
tsk_rows = []
for row in ws3.iter_rows(min_row=6, max_row=200, min_col=1, max_col=1):
    for cell in row:
        v = cell.value
        if v and str(v).strip().startswith("PRY"):
            tsk_rows.append(cell.row)

last_tsk_row = max(tsk_rows) if tsk_rows else 21
TSK_EXTRA = 30
new_last_tsk = last_tsk_row + TSK_EXTRA

TSK_NCOLS = 25
for row_num in range(last_tsk_row + 1, new_last_tsk + 1):
    for col in range(1, TSK_NCOLS + 1):
        c = ws3.cell(row=row_num, column=col)
        c.fill  = fill(CREMA if row_num % 2 == 0 else BLANCO)
        c.font  = Font(name="Arial", size=10, color=GRIS_TEXTO)
        c.alignment = Alignment(
            horizontal="center" if col >= 13 else "left",
            vertical="center")
        c.border = thin_border()

ws3.tables.pop('TAREAS')
new_tsk_tbl = Table(
    displayName="TAREAS",
    ref=f"A5:{get_column_letter(TSK_NCOLS)}{new_last_tsk}"
)
new_tsk_tbl.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9", showRowStripes=True)
ws3.add_table(new_tsk_tbl)
print(f"Updated TAREAS table to A5:{get_column_letter(TSK_NCOLS)}{new_last_tsk}")

# ──────────────────────────────────────────────────────────────────────────────
# Save
# ──────────────────────────────────────────────────────────────────────────────
out = "/home/claude/Matriz_Comunicaciones_AMPA_2026_v2.xlsx"
wb.save(out)
print(f"Saved: {out}")
