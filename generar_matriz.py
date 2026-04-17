from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule, CellIsRule
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.chart import BarChart, Reference
import re

wb = Workbook()

# ─── COLORS ───────────────────────────────────────────────────────────────────
VERDE_SELVA  = "1A3A2A"
VERDE_MEDIO  = "2D5A3D"
VERDE_HOJA   = "3D7A52"
VERDE_LIMA   = "6AB04C"
VERDE_CLARO  = "A8D8A8"
TIERRA       = "8B5E3C"
TIERRA_CLARO = "C4956A"
OCRE         = "D4A853"
OCRE_CLARO   = "F0D080"
CREMA        = "FAF6EE"
CREMA_OSC    = "F0E8D8"
BLANCO       = "FFFFFF"
GRIS_TEXTO   = "4A4A4A"
GRIS_SUAVE   = "E8E0D0"
ROJO         = "C0392B"
AZUL         = "2980B9"
AMARILLO     = "FFF9C4"
VERDE_SUAVE  = "E8F5E9"
ROJO_SUAVE   = "FFEBEE"
NARANJO_SV   = "FFF3E0"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=GRIS_TEXTO, size=10, italic=False, name="Arial"):
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_thin(sides="all"):
    s = Side(style='thin', color="CCCCCC")
    n = Side(style=None)
    t = s if 't' in sides or sides == 'all' else n
    b = s if 'b' in sides or sides == 'all' else n
    l = s if 'l' in sides or sides == 'all' else n
    r = s if 'r' in sides or sides == 'all' else n
    return Border(top=t, bottom=b, left=l, right=r)

def border_medium():
    s = Side(style='medium', color=VERDE_SELVA)
    return Border(top=s, bottom=s, left=s, right=s)

def header_cell(ws, row, col, text, bg=VERDE_SELVA, fg=BLANCO, size=10, bold=True, wrap=False, merge_to=None):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg)
    c.font = font(bold=bold, color=fg, size=size)
    c.alignment = align("center", "center", wrap)
    c.border = border_thin()
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=merge_to[0], end_column=merge_to[1])
    return c

def data_cell(ws, row, col, value="", bg=BLANCO, fg=GRIS_TEXTO, bold=False,
              h="left", wrap=False, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = font(bold=bold, color=fg, size=10, italic=italic)
    c.alignment = align(h, "center", wrap)
    c.border = border_thin()
    return c

def section_title(ws, row, col, text, span_to_col):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=span_to_col)
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(VERDE_MEDIO)
    c.font = font(bold=True, color=BLANCO, size=11)
    c.alignment = align("left", "center")
    c.border = border_thin()
    return c


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: RESUMEN GENERAL
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Resumen General"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A5"

# Title
ws1.merge_cells("A1:N2")
c = ws1["A1"]
c.value = "MATRIZ DE GESTIÓN DE COMUNICACIONES — AMPA 2026"
c.fill = fill(VERDE_SELVA)
c.font = Font(name="Arial", bold=True, size=16, color=BLANCO)
c.alignment = align("center", "center")

ws1.merge_cells("A3:N3")
c = ws1["A3"]
c.value = "⬤  Resumen General de Proyectos  |  Actualiza la tabla de proyectos y el resto de hojas se vinculan automáticamente"
c.fill = fill(VERDE_MEDIO)
c.font = font(italic=True, color=OCRE_CLARO, size=10)
c.alignment = align("center", "center")

ws1.row_dimensions[1].height = 36
ws1.row_dimensions[2].height = 20
ws1.row_dimensions[3].height = 22

# ─── KPI row ──────────────────────────────────────────────────────────────────
row_kpi = 5

kpi_labels = ["PROYECTOS ACTIVOS", "ACTIVIDADES TOTALES", "INDICADORES", "ACTIVIDADES PRÓXIMAS (30d)"]
kpi_cols   = [1, 4, 7, 10]
kpi_widths = [3, 3, 3, 3]

kpi_formulas = [
    '=COUNTIF(Proyectos[Estado],"Activo")',
    '=SUM(Proyectos[N° Actividades])',
    '=SUM(Proyectos[N° Indicadores])',
    '=COUNTIF(Actividades[Estado],"En proceso")',
]
for i, (lbl, col) in enumerate(zip(kpi_labels, kpi_cols)):
    end_col = col + 2
    ws1.merge_cells(start_row=row_kpi, start_column=col, end_row=row_kpi, end_column=end_col)
    c = ws1.cell(row=row_kpi, column=col)
    c.value = lbl
    c.fill = fill(VERDE_HOJA)
    c.font = font(bold=True, color=BLANCO, size=9)
    c.alignment = align("center", "center")

    ws1.merge_cells(start_row=row_kpi+1, start_column=col, end_row=row_kpi+1, end_column=end_col)
    vc = ws1.cell(row=row_kpi+1, column=col)
    vc.value = kpi_formulas[i]
    vc.fill = fill(VERDE_SUAVE)
    vc.font = font(bold=True, color=VERDE_SELVA, size=22)
    vc.alignment = align("center", "center")

ws1.row_dimensions[row_kpi].height = 18
ws1.row_dimensions[row_kpi+1].height = 36

# ─── PROJECTS TABLE ───────────────────────────────────────────────────────────
row_th = 9
headers_pry = ["Código", "Nombre del Proyecto", "Financiador", "Estado",
               "Inicio", "Fin", "% Avance", "N° Actividades", "N° Indicadores",
               "Responsable", "Notas / Observaciones"]

col_widths_pry = [10, 32, 22, 13, 10, 10, 10, 13, 13, 18, 30]

for i, (h, w) in enumerate(zip(headers_pry, col_widths_pry), start=1):
    header_cell(ws1, row_th, i, h)
    ws1.column_dimensions[get_column_letter(i)].width = w

# ─ Sample data ────────────────────────────────────────────────────────────────
proyectos_data = [
    ["PRY-01","Voces del Río","ICCO Cooperación","Activo","Feb-2025","Feb-2026",65,8,5,"Coord. Comunicaciones",""],
    ["PRY-02","Amazonía Viva","Rainforest Foundation","Activo","Ene-2025","Dic-2025",40,6,4,"Coord. Contenidos",""],
    ["PRY-03","Guardianas del Bosque","Global Fund for Women","Activo","Oct-2024","Oct-2025",80,5,3,"Coord. Género",""],
    ["PRY-04","Territorio Común","Oxfam","Activo","Mar-2025","Mar-2026",30,7,5,"Coord. Digital",""],
    ["PRY-05","Agua Viva","WWF Perú","Activo","Nov-2024","Nov-2025",55,4,3,"Directora Prog.",""],
    ["PRY-06","Voz Indígena","Cultural Survival","Activo","Abr-2025","Abr-2026",20,6,4,"Coord. Audiovisual",""],
    ["PRY-07","Semillas Digitales","CAF","Planificación","Jun-2025","Jun-2026",10,7,4,"Coord. Digital",""],
    ["PRY-08","Defensoras","Urgent Action Fund","Activo","Ago-2024","Ago-2025",70,4,3,"Directora Prog.",""],
    ["PRY-09","WLT II","WLT","Activo","Ene-2026","Dic-2026",15,9,6,"Coord. Comunicaciones",""],
    ["PRY-10","CLUA","CLUA","Activo","Ene-2026","Dic-2026",10,8,5,"Coord. Digital",""],
]

dv_estado = DataValidation(type="list", formula1='"Activo,Planificación,Cierre,Suspendido"', showDropDown=False)
ws1.add_data_validation(dv_estado)

for i, row in enumerate(proyectos_data, start=row_th+1):
    bg = CREMA if i % 2 == 0 else BLANCO
    for j, val in enumerate(row, start=1):
        c = data_cell(ws1, i, j, val, bg=bg)
        if j == 7:
            c.number_format = '0"%"'
        if j == 4:
            dv_estado.add(c)

# ─ Table ──────────────────────────────────────────────────────────────────────
last_row_pry = row_th + len(proyectos_data)
table_pry = Table(displayName="Proyectos",
                  ref=f"A{row_th}:{get_column_letter(len(headers_pry))}{last_row_pry}")
style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
table_pry.tableStyleInfo = style
ws1.add_table(table_pry)

# ─ Conditional formatting for Estado ──────────────────────────────────────────
ws1.conditional_formatting.add(
    f"D{row_th+1}:D{last_row_pry+20}",
    CellIsRule(operator='equal', formula=['"Activo"'], fill=fill("E8F5E9"),
               font=Font(color="2E7D32", bold=True)))
ws1.conditional_formatting.add(
    f"D{row_th+1}:D{last_row_pry+20}",
    CellIsRule(operator='equal', formula=['"Planificación"'], fill=fill("FFF8E1"),
               font=Font(color="F57F17", bold=True)))
ws1.conditional_formatting.add(
    f"D{row_th+1}:D{last_row_pry+20}",
    CellIsRule(operator='equal', formula=['"Cierre"'], fill=fill("FCE4EC"),
               font=Font(color="C62828", bold=True)))

# ─ Instructions box ───────────────────────────────────────────────────────────
ws1.merge_cells(f"A{last_row_pry+2}:K{last_row_pry+5}")
c = ws1.cell(row=last_row_pry+2, column=1,
    value="📌  CÓMO AGREGAR UN NUEVO PROYECTO: Haz clic en cualquier celda de la tabla → Ve a la última fila → "
          "Escribe en la siguiente fila vacía debajo de la tabla (se expandirá automáticamente). "
          "Luego ve a la hoja 'Actividades por Proyecto' y agrega las actividades del nuevo proyecto. "
          "Los KPIs de esta hoja se actualizan solos.")
c.fill = fill(OCRE_CLARO)
c.font = font(italic=True, color=VERDE_SELVA, size=9)
c.alignment = align("left", "top", wrap=True)
c.border = border_thin()
ws1.row_dimensions[last_row_pry+2].height = 60


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: ACTIVIDADES POR PROYECTO
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Actividades por Proyecto")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A5"

ws2.merge_cells("A1:P2")
c = ws2["A1"]
c.value = "ACTIVIDADES COMUNICACIONALES POR PROYECTO"
c.fill = fill(VERDE_SELVA)
c.font = Font(name="Arial", bold=True, size=15, color=BLANCO)
c.alignment = align("center", "center")
ws2.row_dimensions[1].height = 32

ws2.merge_cells("A3:P3")
c = ws2["A3"]
c.value = "El % Avance de cada actividad se calcula automáticamente desde la hoja TAREAS (suma de tareas completadas / total de tareas)"
c.fill = fill(VERDE_MEDIO)
c.font = font(italic=True, color=OCRE_CLARO, size=9)
c.alignment = align("center", "center")
ws2.row_dimensions[3].height = 18

act_headers = ["Proyecto", "Componente / Objetivo", "Actividad / Producto",
               "Tipo", "Resultado esperado", "Indicador vinculado",
               "Responsable", "Co-responsable", "Fecha Inicio", "Fecha Entrega",
               "Prioridad", "Estado", "% Avance (auto)", "Tareas totales",
               "Tareas completas", "Observaciones"]
act_widths  = [12, 24, 36, 14, 30, 26, 16, 16, 12, 12, 10, 12, 13, 12, 13, 32]

row_ath = 5
for i, (h, w) in enumerate(zip(act_headers, act_widths), start=1):
    header_cell(ws2, row_ath, i, h)
    ws2.column_dimensions[get_column_letter(i)].width = w

actividades_data = [
    ["PRY-01","Comp. 2 — Difusión","Producción de documental corto sobre comunidades ribeñas","Audiovisual","1 video documental de 15 min publicado","IND-2.1 — N. materiales audiovisuales","Coord. Audiovisual","","Mar-2025","Jun-2025","Alta","En proceso","","","","En edición, falta locución"],
    ["PRY-01","Comp. 3 — Incidencia","Campaña en RRSS sobre derechos territoriales","Redes Sociales","12 piezas publicadas, alcance 5.000 personas","IND-3.2 — Alcance RRSS","Coord. Digital","","Abr-2025","May-2025","Alta","En proceso","","","","Revisión de mensajes clave pendiente"],
    ["PRY-01","Comp. 1 — Fortalecimiento","Boletín informativo para comunidades (x3)","Difusión","3 boletines impresos y digitales distribuidos","IND-1.4 — N. boletines","Directora Prog.","","Feb-2025","Ago-2025","Media","En proceso","","","","1 de 3 boletines completado"],
    ["PRY-02","Comp. 2 — Sensibilización","Serie de podcasts sobre biodiversidad amazónica","Audiovisual","6 episodios publicados en plataformas digitales","IND-2.3 — N. podcasts producidos","Coord. Contenidos","","Ene-2025","Jul-2025","Media","En proceso","","","","4 episodios listos, 2 en grabación"],
    ["PRY-02","Comp. 1 — Educación","Materiales educativos para escuelas rurales","Publicación","500 cartillas impresas distribuidas","IND-1.1 — N. cartillas","Coord. Ed. Ambiental","","Feb-2025","May-2025","Alta","Retrasado","","","","URGENTE: diseño pendiente de aprobación"],
    ["PRY-03","Comp. 1 — Género","Video testimonial de guardianas del bosque","Audiovisual","1 video de 8 min para difusión nacional","IND-1.2 — Materiales con enfoque de género","Coord. Género","","Nov-2024","Feb-2025","Alta","Completado","","","","Publicado en web y RRSS"],
    ["PRY-04","Comp. 3 — Incidencia","Campaña de incidencia política en redes","Campaña","Campaña activa por 30 días, 10.000 impactos","IND-3.1 — Impacto campaña digital","Coord. Digital","","Abr-2025","Jun-2025","Alta","Pendiente","","","","Requiere aprobación de mensajes con aliados"],
    ["PRY-05","Comp. 2 — Comunicación","Mapa interactivo de fuentes de agua amenazadas","Publicación","1 infografía digital publicada","IND-2.4 — Herramientas de comunicación","Coord. SIG + Comm.","","Feb-2025","Abr-2025","Media","Retrasado","","","","Retraso por datos faltantes de campo"],
    ["PRY-06","Comp. 2 — Revitalización","Producción de material audiovisual en lengua shipibo","Audiovisual","3 videos cortos en lengua originaria","IND-2.2 — Mat. en lenguas indígenas","Coord. Audiovisual","","May-2025","Sep-2025","Alta","Pendiente","","","","Coordinación con comunicadores indígenas locales"],
    ["PRY-07","Comp. 1 — Formación","Taller de comunicación digital para jóvenes amazónicos","Evento","40 jóvenes capacitados, 1 informe final","IND-1.1 — N. participantes talleres","Coord. Digital","","Jul-2025","Ago-2025","Media","Pendiente","","","","Confirmar sedes con comunidades"],
    ["PRY-08","Comp. 2 — Protección","Protocolo de comunicación segura para defensoras","Publicación","1 guía impresa y digital distribuida","IND-2.1 — Herramientas de protección","Directora Prog.","","Nov-2024","Mar-2025","Alta","Completado","","","","Distribuida en 3 comunidades"],
    ["PRY-09","Obj. 1 — Fortalecimiento","Fortalecimiento capacidades ABIOFORP — redes sociales turísticas","Capacitación","2 talleres ejecutados, informe final","IND-1.1","Coord. Comunicaciones","Coord. Digital","Ene-2026","Jun-2026","Alta","En proceso","","","","En coordinación con AMPA y CxN"],
    ["PRY-09","Obj. 1 — Fortalecimiento","Communication toolkit sociobioeconomía","Publicación","Kit de materiales terminado","IND-1.2","Coord. Comunicaciones","Coord. Audiovisual","Mar-2026","Jul-2026","Media","Pendiente","","","",""],
    ["PRY-10","Obj. 1 — CLUA","Escuela Amazonía Que Late — actualización plataforma","Digital","Plataforma actualizada con info bioeconomía","IND-1.1","Coord. Digital","Coord. Contenidos","Ene-2026","May-2026","Alta","En proceso","","","","Diagnóstico de necesidades en curso"],
    ["PRY-10","Obj. 1 — CLUA","Conservamos Festival — planificación y ejecución","Evento","Festival ejecutado, 1000+ asistentes","IND-1.4","Coord. Digital","Directora Prog.","Abr-2026","Nov-2026","Alta","Planificación","","","","Pendiente cerrar espacio con municipalidad"],
]

dv_estado2 = DataValidation(type="list",
    formula1='"Pendiente,En proceso,Completado,Retrasado,Cancelado"', showDropDown=False)
dv_prior = DataValidation(type="list", formula1='"Alta,Media,Baja"', showDropDown=False)
dv_tipo  = DataValidation(type="list",
    formula1='"Audiovisual,Redes Sociales,Prensa,Publicación,Evento,Difusión,Campaña,Capacitación,Digital,Otro"',
    showDropDown=False)
ws2.add_data_validation(dv_estado2)
ws2.add_data_validation(dv_prior)
ws2.add_data_validation(dv_tipo)

for i, row in enumerate(actividades_data, start=row_ath+1):
    bg = CREMA if i % 2 == 0 else BLANCO
    for j, val in enumerate(row, start=1):
        c = data_cell(ws2, i, j, val, bg=bg, wrap=(j in [3,5,6,16]))
        if j == 12:
            dv_estado2.add(c)
        if j == 11:
            dv_prior.add(c)
        if j == 4:
            dv_tipo.add(c)
    # Auto-calculate % Avance from TAREAS sheet (col 13)
    act_id = i - row_ath  # activity number (1-based)
    r = i
    ws2.cell(row=r, column=13).value = (
        f'=IFERROR(COUNTIFS(TAREAS[ID Actividad],A{r},TAREAS[Estado Tarea],"Completada")'
        f'/COUNTIF(TAREAS[ID Actividad],A{r}&"_"&ROW()-{row_ath}),"")'
    )
    ws2.cell(row=r, column=14).value = f'=COUNTIF(TAREAS[ID Actividad],A{r}&"_{act_id}")'
    ws2.cell(row=r, column=15).value = (
        f'=COUNTIFS(TAREAS[ID Actividad],A{r}&"_{act_id}",TAREAS[Estado Tarea],"Completada")'
    )
    ws2.cell(row=r, column=13).number_format = '0%'

last_row_act = row_ath + len(actividades_data)
table_act = Table(displayName="Actividades",
                  ref=f"A{row_ath}:{get_column_letter(len(act_headers))}{last_row_act}")
table_act.tableStyleInfo = TableStyleInfo(name="TableStyleMedium7", showRowStripes=True)
ws2.add_table(table_act)

# Conditional formatting
for estado, bg_c, fg_c in [("En proceso", NARANJO_SV, "BF360C"),
                            ("Completado", VERDE_SUAVE, "1B5E20"),
                            ("Retrasado",  ROJO_SUAVE,  "B71C1C"),
                            ("Pendiente",  "F5F5F5",    "424242")]:
    ws2.conditional_formatting.add(
        f"L{row_ath+1}:L{last_row_act+50}",
        CellIsRule(operator='equal', formula=[f'"{estado}"'],
                   fill=fill(bg_c), font=Font(color=fg_c, bold=True)))

ws2.merge_cells(f"A{last_row_act+2}:P{last_row_act+4}")
c = ws2.cell(row=last_row_act+2, column=1,
    value="📌  Para agregar una nueva actividad: escribe en la fila siguiente al final de la tabla. "
          "Luego ve a la hoja TAREAS y agrega las tareas correspondientes con el código de proyecto y número de actividad. "
          "El % Avance se actualizará automáticamente.")
c.fill = fill(OCRE_CLARO)
c.font = font(italic=True, color=VERDE_SELVA, size=9)
c.alignment = align("left", "top", wrap=True)
c.border = border_thin()
ws2.row_dimensions[last_row_act+2].height = 50


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: TAREAS
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("TAREAS")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A6"

ws3.merge_cells("A1:R2")
c = ws3["A1"]
c.value = "TAREAS POR ACTIVIDAD — Plan de Trabajo Detallado"
c.fill = fill(VERDE_SELVA)
c.font = Font(name="Arial", bold=True, size=15, color=BLANCO)
c.alignment = align("center", "center")
ws3.row_dimensions[1].height = 32

ws3.merge_cells("A3:R3")
c = ws3["A3"]
c.value = ("ID Actividad = Código Proyecto + '_' + N° de fila de la actividad en 'Actividades por Proyecto' "
           "(ej: PRY-01_1). El avance vincula automáticamente con 'Actividades por Proyecto'.")
c.fill = fill(VERDE_MEDIO)
c.font = font(italic=True, color=OCRE_CLARO, size=9)
c.alignment = align("center", "center")
ws3.row_dimensions[3].height = 22

task_headers = ["ID Actividad", "Proyecto", "Actividad", "N° Tarea", "Descripción de Tarea",
                "Responsable", "Co-responsable", "Estado Tarea", "Prioridad",
                "Fecha Inicio", "Fecha Límite", "Fecha Completada",
                "Ene", "Feb", "Mar", "Abr", "May", "Jun",
                "Jul", "Ago", "Sep", "Oct", "Nov", "Dic",
                "Comentarios / Notas"]
task_widths  = [14, 10, 30, 8, 36, 16, 16, 13, 10,
                12, 12, 14,
                5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5,
                32]

row_tth = 5
for i, (h, w) in enumerate(zip(task_headers, task_widths), start=1):
    header_cell(ws3, row_tth, i, h)
    ws3.column_dimensions[get_column_letter(i)].width = w

# ─── Month headers sub-row ────────────────────────────────────────────────────
ws3.merge_cells(f"M4:X4")
c = ws3.cell(row=4, column=13, value="CRONOGRAMA MENSUAL (marcar con X o ✓)")
c.fill = fill(VERDE_HOJA)
c.font = font(bold=True, color=BLANCO, size=9)
c.alignment = align("center", "center")
ws3.row_dimensions[4].height = 18

# ─── Sample tasks ─────────────────────────────────────────────────────────────
tareas_data = [
    ["PRY-01_1","PRY-01","Producción documental corto",1,"Elaborar guión del documental","Coord. Audiovisual","","Completada","Alta","Mar-2025","Mar-2025","Mar-2025","","","X","","","","","","","","","","Aprobado por dirección"],
    ["PRY-01_1","PRY-01","Producción documental corto",2,"Grabación en comunidades","Coord. Audiovisual","Fotografa","Completada","Alta","Mar-2025","Abr-2025","Abr-2025","","","","X","","","","","","","","","Filmación completada en 3 comunidades"],
    ["PRY-01_1","PRY-01","Producción documental corto",3,"Edición y posproducción","Coord. Audiovisual","","En proceso","Alta","May-2025","Jun-2025","","","","","","X","X","","","","","","","Falta locución en off"],
    ["PRY-01_2","PRY-01","Campaña RRSS derechos territoriales",1,"Definir mensajes clave con aliados","Coord. Digital","Directora Prog.","En proceso","Alta","Abr-2025","Abr-2025","","","","","X","","","","","","","","","Revisión pendiente con 3 organizaciones"],
    ["PRY-01_2","PRY-01","Campaña RRSS derechos territoriales",2,"Diseñar 12 piezas gráficas","Coord. Digital","","Pendiente","Alta","Abr-2025","May-2025","","","","","X","X","","","","","","","","Espera aprobación de mensajes"],
    ["PRY-01_3","PRY-01","Boletín informativo (x3)",1,"Redactar boletín N°1","Directora Prog.","","Completada","Media","Feb-2025","Feb-2025","Feb-2025","","X","","","","","","","","","","","Distribuido a 200 contactos"],
    ["PRY-01_3","PRY-01","Boletín informativo (x3)",2,"Redactar boletín N°2","Directora Prog.","","Pendiente","Media","May-2025","May-2025","","","","","","X","","","","","","","",""],
    ["PRY-09_1","PRY-09","Fortalecimiento capacidades ABIOFORP",1,"Diagnóstico de necesidades de alfabetización digital","Coord. Comunicaciones","Coord. Digital","Completada","Alta","Ene-2026","Feb-2026","Feb-2026","X","X","","","","","","","","","","","Coordinado con AMPA y CxN. Reunión 6 abril."],
    ["PRY-09_1","PRY-09","Fortalecimiento capacidades ABIOFORP",2,"1° Taller de alfabetización digital (Jornada 1)","Coord. Comunicaciones","Coord. Digital","En proceso","Alta","Feb-2026","Mar-2026","","","X","X","","","","","","","","","","Compartir matriz con CxN para revisar"],
    ["PRY-09_1","PRY-09","Fortalecimiento capacidades ABIOFORP",3,"1° Taller de alfabetización digital (Jornada 2)","Coord. Comunicaciones","Coord. Digital","Pendiente","Alta","Mar-2026","Abr-2026","","","","X","X","","","","","","","","",""],
    ["PRY-09_1","PRY-09","Fortalecimiento capacidades ABIOFORP",4,"2° Taller — Estrategia Reporteros del Bosque","Coord. Comunicaciones","","Pendiente","Alta","May-2026","Jun-2026","","","","","","X","X","","","","","","","Linkear desde otras plataformas (CxN)"],
    ["PRY-10_1","PRY-10","Escuela Amazonía Que Late — actualización",1,"Diagnóstico de necesidades de actualización plataforma","Coord. Digital","Coord. Contenidos","En proceso","Alta","Ene-2026","Feb-2026","","X","X","","","","","","","","","","","Evaluar desde punto de vista usuario y administración"],
    ["PRY-10_1","PRY-10","Escuela Amazonía Que Late — actualización",2,"Revisar y actualizar matrices existentes","Coord. Digital","Coord. Contenidos","Pendiente","Alta","Feb-2026","Mar-2026","","","X","X","","","","","","","","","","Compartir matriz con CxN para revisar y actualizar"],
    ["PRY-10_3","PRY-10","Conservamos Festival — planificación",1,"Reuniones con auspiciadores y aliados","Coord. Digital","Directora Prog.","En proceso","Alta","Ene-2026","May-2026","","X","X","X","X","X","","","","","","","","Participación conjunta AMPA y CxN"],
    ["PRY-10_3","PRY-10","Conservamos Festival — planificación",2,"TdR Quantico","Coord. Digital","","En proceso","Alta","Feb-2026","Mar-2026","","","X","X","","","","","","","","","",""],
    ["PRY-10_3","PRY-10","Conservamos Festival — planificación",3,"Ejecución del Festival","Coord. Digital","Directora Prog.","Pendiente","Alta","Nov-2026","Nov-2026","","","","","","","","","","","","X","","21 de noviembre"],
]

dv_estado3 = DataValidation(type="list",
    formula1='"Pendiente,En proceso,Completada,Bloqueada,Cancelada"', showDropDown=False)
dv_prior3 = DataValidation(type="list", formula1='"Alta,Media,Baja"', showDropDown=False)
ws3.add_data_validation(dv_estado3)
ws3.add_data_validation(dv_prior3)

for i, row in enumerate(tareas_data, start=row_tth+1):
    bg = CREMA if i % 2 == 0 else BLANCO
    for j, val in enumerate(row, start=1):
        is_month = 13 <= j <= 24
        c = data_cell(ws3, i, j, val, bg=bg,
                      h="center" if (is_month or j in [4,8,9]) else "left",
                      wrap=(j in [5,25]))
        if j == 8:
            dv_estado3.add(c)
        if j == 9:
            dv_prior3.add(c)
        if is_month and val:
            c.fill = fill(VERDE_CLARO)
            c.font = font(bold=True, color=VERDE_SELVA)

    ws3.row_dimensions[i].height = 20

last_row_tsk = row_tth + len(tareas_data)
table_tsk = Table(displayName="TAREAS",
                  ref=f"A{row_tth}:{get_column_letter(len(task_headers))}{last_row_tsk}")
table_tsk.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
ws3.add_table(table_tsk)

# Conditional formatting for Estado Tarea
for estado, bg_c, fg_c in [("En proceso",  NARANJO_SV, "BF360C"),
                            ("Completada",  VERDE_SUAVE,"1B5E20"),
                            ("Bloqueada",   ROJO_SUAVE, "B71C1C"),
                            ("Pendiente",   "F5F5F5",   "424242"),
                            ("Cancelada",   "EEEEEE",   "757575")]:
    ws3.conditional_formatting.add(
        f"H{row_tth+1}:H{last_row_tsk+100}",
        CellIsRule(operator='equal', formula=[f'"{estado}"'],
                   fill=fill(bg_c), font=Font(color=fg_c, bold=True)))

ws3.merge_cells(f"A{last_row_tsk+2}:Y{last_row_tsk+5}")
c = ws3.cell(row=last_row_tsk+2, column=1,
    value="📌  CÓMO AGREGAR TAREAS: Escribe en la fila siguiente al final de la tabla. "
          "ID Actividad = [Código Proyecto]_[N° fila actividad], ej: PRY-01_3. "
          "Marca con X los meses en el cronograma. "
          "El % avance en 'Actividades por Proyecto' se calculará según cuántas tareas estén 'Completada'.")
c.fill = fill(OCRE_CLARO)
c.font = font(italic=True, color=VERDE_SELVA, size=9)
c.alignment = align("left", "top", wrap=True)
c.border = border_thin()
ws3.row_dimensions[last_row_tsk+2].height = 60


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: INDICADORES DE COMUNICACIÓN
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Indicadores Comunicación")
ws4.sheet_view.showGridLines = False
ws4.freeze_panes = "A6"

ws4.merge_cells("A1:N2")
c = ws4["A1"]
c.value = "INDICADORES DE COMUNICACIÓN — Seguimiento por Proyecto"
c.fill = fill(VERDE_SELVA)
c.font = Font(name="Arial", bold=True, size=15, color=BLANCO)
c.alignment = align("center", "center")
ws4.row_dimensions[1].height = 32

ws4.merge_cells("A3:N3")
c = ws4["A3"]
c.value = "Registra aquí los indicadores de cada proyecto. Actualiza 'Avance Actual' para ver el % logrado automáticamente."
c.fill = fill(VERDE_MEDIO)
c.font = font(italic=True, color=OCRE_CLARO, size=9)
c.alignment = align("center", "center")

ind_headers = ["Proyecto", "Objetivo / Componente", "Nombre del Indicador",
               "Medio de Verificación", "Línea Base", "Meta Anual",
               "Avance T1 (Ene-Mar)", "Avance T2 (Abr-Jun)",
               "Avance T3 (Jul-Sep)", "Avance T4 (Oct-Dic)",
               "Total Acumulado", "% Logro", "Fuente / Evidencia", "Notas"]
ind_widths   = [12, 20, 36, 22, 10, 10, 14, 14, 14, 14, 14, 10, 22, 28]

row_ith = 5
for i, (h, w) in enumerate(zip(ind_headers, ind_widths), start=1):
    header_cell(ws4, row_ith, i, h)
    ws4.column_dimensions[get_column_letter(i)].width = w

indicadores_data = [
    ["PRY-01","Comp. 2 — Difusión","N° de materiales audiovisuales producidos y publicados","Video publicado en web/RRSS",0,4,2,0,0,0,"","","",""],
    ["PRY-01","Comp. 3 — Incidencia","Alcance de campañas en RRSS (personas impactadas)","Estadísticas plataformas",0,15000,7200,0,0,0,"","","",""],
    ["PRY-01","Comp. 1 — Fortalecimiento","N° de boletines informativos distribuidos","Lista de distribución",0,3,1,0,0,0,"","","",""],
    ["PRY-02","Comp. 2 — Sensibilización","N° de episodios de podcast producidos","Spotify/SoundCloud links",0,6,4,0,0,0,"","","",""],
    ["PRY-02","Comp. 1 — Educación","N° de cartillas educativas distribuidas en escuelas","Actas de entrega",0,500,0,0,0,0,"","","",""],
    ["PRY-03","Comp. 1 — Género","N° de materiales comunicacionales con enfoque de género","Registro de materiales",0,5,4,0,0,0,"","","",""],
    ["PRY-03","Comp. 2 — Visibilización","N° de notas publicadas en medios regionales","Links/clippings",0,3,5,0,0,0,"","","","Superó meta"],
    ["PRY-04","Comp. 3 — Incidencia","Impacto campaña digital de incidencia (impresiones)","Facebook/Instagram Insights",0,10000,0,0,0,0,"","","",""],
    ["PRY-05","Comp. 2 — Comunicación","N° de herramientas de comunicación ambiental creadas","Registro publicaciones",0,2,0,0,0,0,"","","",""],
    ["PRY-06","Comp. 2 — Revitalización","N° de materiales producidos en lenguas indígenas","Registro audiovisual",0,3,0,0,0,0,"","","",""],
    ["PRY-07","Comp. 1 — Formación","N° de jóvenes capacitados en comunicación digital","Lista de participantes",0,40,0,0,0,0,"","","",""],
    ["PRY-08","Comp. 2 — Protección","N° de guías de comunicación segura distribuidas","Constancias de entrega",0,200,200,0,0,0,"","","","Meta cumplida"],
    ["PRY-08","Comp. 1 — Visibilidad","Alcance de campaña de visibilidad de defensoras","Instagram/Twitter Insights",0,15000,21000,0,0,0,"","","","Superó meta en 40%"],
    ["PRY-09","Obj. 1 — Fortalecimiento","N° de talleres de capacitación ejecutados","Informes de taller",0,4,1,0,0,0,"","","",""],
    ["PRY-09","Obj. 1 — Fortalecimiento","N° de materiales del toolkit producidos","Registro de materiales",0,1,0,0,0,0,"","","",""],
    ["PRY-10","Obj. 1 — CLUA","N° de actualizaciones implementadas en plataforma","Registro web",0,5,0,0,0,0,"","","",""],
    ["PRY-10","Obj. 1 — CLUA","N° de asistentes al Conservamos Festival","Lista de asistentes",0,1000,0,0,0,0,"","","",""],
]

for i, row in enumerate(indicadores_data, start=row_ith+1):
    bg = CREMA if i % 2 == 0 else BLANCO
    for j, val in enumerate(row, start=1):
        c = data_cell(ws4, i, j, val, bg=bg, wrap=(j in [3,4,13,14]))
    # Formula: Total = T1+T2+T3+T4
    r = i
    ws4.cell(row=r, column=11).value = f'=SUM(G{r}:J{r})'
    # % Logro = Total / Meta
    ws4.cell(row=r, column=12).value = f'=IFERROR(K{r}/F{r},"")'
    ws4.cell(row=r, column=12).number_format = '0%'

last_row_ind = row_ith + len(indicadores_data)
table_ind = Table(displayName="Indicadores",
                  ref=f"A{row_ith}:{get_column_letter(len(ind_headers))}{last_row_ind}")
table_ind.tableStyleInfo = TableStyleInfo(name="TableStyleMedium3", showRowStripes=True)
ws4.add_table(table_ind)

# Color scale on % Logro
ws4.conditional_formatting.add(
    f"L{row_ith+1}:L{last_row_ind+50}",
    ColorScaleRule(start_type='num', start_value=0, start_color='FFEBEE',
                   mid_type='num', mid_value=0.5, mid_color='FFF9C4',
                   end_type='num', end_value=1, end_color='E8F5E9'))

ws4.merge_cells(f"A{last_row_ind+2}:N{last_row_ind+4}")
c = ws4.cell(row=last_row_ind+2, column=1,
    value="📌  CÓMO AGREGAR INDICADORES: Escribe en la fila siguiente a la tabla. Actualiza los avances por trimestre (T1-T4). "
          "El Total Acumulado y % Logro se calculan automáticamente. Puedes filtrar por proyecto en la columna A.")
c.fill = fill(OCRE_CLARO)
c.font = font(italic=True, color=VERDE_SELVA, size=9)
c.alignment = align("left", "top", wrap=True)
c.border = border_thin()
ws4.row_dimensions[last_row_ind+2].height = 50


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5: CRONOGRAMA
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Cronograma")
ws5.sheet_view.showGridLines = False
ws5.freeze_panes = "D5"

ws5.merge_cells("A1:P2")
c = ws5["A1"]
c.value = "CRONOGRAMA CONSOLIDADO 2026 — Actividades Comunicacionales"
c.fill = fill(VERDE_SELVA)
c.font = Font(name="Arial", bold=True, size=15, color=BLANCO)
c.alignment = align("center", "center")
ws5.row_dimensions[1].height = 32

ws5.merge_cells("A3:P3")
c = ws5["A3"]
c.value = "Vista mensual. Las celdas de mes se marcan automáticamente desde TAREAS. También puedes marcar directamente. El mes actual se resalta en amarillo."
c.fill = fill(VERDE_MEDIO)
c.font = font(italic=True, color=OCRE_CLARO, size=9)
c.alignment = align("center", "center")
ws5.row_dimensions[3].height = 18

meses = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
crono_headers = ["Proyecto","Actividad / Producto","Responsable"] + meses
crono_widths  = [12, 38, 16] + [5]*12

row_cth = 4
for i, (h, w) in enumerate(zip(crono_headers, crono_widths), start=1):
    bg = VERDE_SELVA if i <= 3 else VERDE_HOJA
    header_cell(ws5, row_cth, i, h, bg=bg)
    ws5.column_dimensions[get_column_letter(i)].width = w

crono_rows = [
    ["PRY-01","Producción documental corto","Coord. Audiovisual","","","X","X","X","","","","","","",""],
    ["PRY-01","Campaña RRSS derechos territoriales","Coord. Digital","","","","X","X","","","","","","",""],
    ["PRY-01","Boletín informativo (x3)","Directora Prog.","","X","","","X","","","X","","","",""],
    ["PRY-02","Serie podcasts biodiversidad","Coord. Contenidos","X","X","X","X","X","X","X","","","","",""],
    ["PRY-02","Materiales educativos escuelas","Coord. Ed. Ambiental","","X","X","X","X","","","","","","",""],
    ["PRY-03","Video testimonial guardianas","Coord. Género","","","","","","","","","","","",""],
    ["PRY-04","Cobertura talleres y eventos","Fotografa","","","X","X","X","X","X","X","X","X","",""],
    ["PRY-04","Campaña incidencia política","Coord. Digital","","","","X","X","X","","","","","",""],
    ["PRY-05","Mapa interactivo fuentes agua","Coord. SIG","","X","X","X","","","","","","","",""],
    ["PRY-05","Informe semestral WWF","Directora Prog.","","","","","","X","X","","","","",""],
    ["PRY-06","Videos en lengua shipibo","Coord. Audiovisual","","","","","X","X","X","X","X","","",""],
    ["PRY-07","Taller comunicación digital jóvenes","Coord. Digital","","","","","","","X","X","","","",""],
    ["PRY-08","Campaña #DefensorasAmazonia","Coord. Digital","","","X","X","","","","","","","",""],
    ["PRY-09","Talleres alfabetización digital","Coord. Comunicaciones","X","X","X","X","X","X","","","","","",""],
    ["PRY-09","Communication toolkit","Coord. Comunicaciones","","","X","X","X","X","X","","","","",""],
    ["PRY-10","Actualización Escuela AQL","Coord. Digital","X","X","X","X","X","","","","","","",""],
    ["PRY-10","Conservamos Festival","Coord. Digital","X","X","X","X","X","X","X","X","X","X","X",""],
]

for i, row in enumerate(crono_rows, start=row_cth+1):
    bg = CREMA if i % 2 == 0 else BLANCO
    for j, val in enumerate(row, start=1):
        c = data_cell(ws5, i, j, val, bg=bg, h="center" if j > 3 else "left",
                      wrap=(j==2))
        if j > 3 and val == "X":
            c.fill = fill(VERDE_LIMA)
            c.font = font(bold=True, color=VERDE_SELVA)
            c.alignment = align("center","center")
    ws5.row_dimensions[i].height = 20

last_row_cro = row_cth + len(crono_rows)
table_cro = Table(displayName="Cronograma",
                  ref=f"A{row_cth}:{get_column_letter(len(crono_headers))}{last_row_cro}")
table_cro.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
ws5.add_table(table_cro)

ws5.merge_cells(f"A{last_row_cro+2}:P{last_row_cro+4}")
c = ws5.cell(row=last_row_cro+2, column=1,
    value="📌  Marca con X el mes en que se ejecuta cada actividad. Para agregar una nueva actividad, escribe en la siguiente fila vacía de la tabla.")
c.fill = fill(OCRE_CLARO)
c.font = font(italic=True, color=VERDE_SELVA, size=9)
c.alignment = align("left", "top", wrap=True)
c.border = border_thin()
ws5.row_dimensions[last_row_cro+2].height = 40


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6: COMUNICACIÓN INSTITUCIONAL
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Comunicación Institucional")
ws6.sheet_view.showGridLines = False
ws6.freeze_panes = "A6"

ws6.merge_cells("A1:M2")
c = ws6["A1"]
c.value = "COMUNICACIÓN INSTITUCIONAL — AMPA"
c.fill = fill(VERDE_SELVA)
c.font = Font(name="Arial", bold=True, size=15, color=BLANCO)
c.alignment = align("center", "center")
ws6.row_dimensions[1].height = 32

ws6.merge_cells("A3:M3")
c = ws6["A3"]
c.value = ("Diferencia clave: Lo institucional es la voz de AMPA como organización (redes, web, vocería). "
           "Los proyectos tienen sus propias comunicaciones. Esta hoja gestiona ambas capas por separado.")
c.fill = fill(VERDE_MEDIO)
c.font = font(italic=True, color=OCRE_CLARO, size=9)
c.alignment = align("center", "center")
ws6.row_dimensions[3].height = 22

inst_headers = ["Bloque / Categoría", "Acción / Producto", "Descripción",
                "Frecuencia", "Responsable", "Canal / Plataforma",
                "Estado", "Fecha / Período", "Ene","Feb","Mar","Abr","May",
                "Jun","Jul","Ago","Sep","Oct","Nov","Dic","Observaciones"]
inst_widths  = [20, 30, 36, 12, 16, 18, 12, 14, 5,5,5,5,5,5,5,5,5,5,5,5, 30]

row_inth = 5
for i, (h, w) in enumerate(zip(inst_headers, inst_widths), start=1):
    header_cell(ws6, row_inth, i, h)
    ws6.column_dimensions[get_column_letter(i)].width = w

institucional_data = [
    # Redes Sociales Institucionales
    ["Redes Sociales Institucionales","Gestión de Instagram AMPA","Publicaciones orgánicas sobre trabajo institucional y proyectos","Semanal","Coord. Digital","Instagram","Activo","Todo el año","X","X","X","X","X","X","X","X","X","X","X","X","3-4 posts por semana"],
    ["Redes Sociales Institucionales","Gestión de Facebook AMPA","Publicaciones y gestión de comunidad","Semanal","Coord. Digital","Facebook","Activo","Todo el año","X","X","X","X","X","X","X","X","X","X","X","X",""],
    ["Redes Sociales Institucionales","Twitter/X Institucional","Comunicados, posicionamiento y articulación con aliados","Diario","Coord. Digital","Twitter/X","Activo","Todo el año","X","X","X","X","X","X","X","X","X","X","X","X",""],
    ["Redes Sociales Institucionales","TikTok institucional","Contenido corto sobre conservación y comunidades","Bisemanal","Coord. Digital","TikTok","En proceso","Todo el año","","","X","X","X","X","X","X","X","X","X","X","En exploración"],
    # Boletines y Comunicados
    ["Boletines y Comunicados","Boletín mensual AMPA","Boletín de actividades y alertas territoriales por WhatsApp y correo","Mensual","Directora Prog.","Email / WhatsApp","Activo","Mensual","X","X","X","X","X","X","X","X","X","X","X","X","Distribución a ~500 contactos"],
    ["Boletines y Comunicados","Nota de prensa institucional","Notas sobre hitos institucionales o posicionamiento público","Según evento","Coord. Prensa","Medios","Según evento","","","","","","","","","","","","","Redactar según hitos clave"],
    ["Boletines y Comunicados","Comunicado de posición","Pronunciamiento ante temas de agenda ambiental o territorial","Según evento","Directora Prog.","Web + RRSS","Según evento","","","","","","","","","","","","","Previa aprobación Dirección Ejecutiva"],
    # Sitio Web
    ["Sitio Web","Actualización de noticias web","Publicación de noticias, eventos y logros institucionales en web","Quincenal","Coord. Digital","Web AMPA","Activo","Quincenal","X","X","X","X","X","X","X","X","X","X","X","X",""],
    ["Sitio Web","Actualización de proyectos en web","Mantener vigentes los perfiles de proyectos activos","Trimestral","Coord. Digital","Web AMPA","Activo","Trimestral","","","X","","","X","","","X","","","X",""],
    # Materiales Institucionales
    ["Materiales Institucionales","Informe anual de actividades","Reporte público anual para aliados, financiadores y comunidad","Anual","Directora Prog.","Impreso + PDF","Planificado","Dic-2026","","","","","","","","","","","","X","Diseño en Q3"],
    ["Materiales Institucionales","Brochure institucional actualizado","Presentación institucional actualizada con proyectos 2026","Anual","Coord. Comunicaciones","PDF + Impreso","Planificado","Mar-2026","","","X","","","","","","","","","",""],
    ["Materiales Institucionales","Presentación institucional PPT","Deck para reuniones con financiadores y aliados","Semestral","Directora Prog.","PowerPoint","Activo","Mar/Sep","","","X","","","","","","X","","","",""],
    # Eventos Institucionales
    ["Eventos Institucionales","Participación en espacios de articulación","Asistencia a eventos de redes y coaliciones","Según agenda","Directora Prog.","Presencial","Activo","Según agenda","X","X","X","X","X","X","X","X","X","X","X","X","Priorizar eventos estratégicos"],
    ["Eventos Institucionales","Organización de eventos propios AMPA","Talleres, foros o eventos convocados por AMPA","Según plan","Coord. Comunicaciones","Presencial","Planificado","","","","","","X","","","X","","","","",""],
    # Monitoreo Institucional
    ["Monitoreo Institucional","Reporte interno mensual","Reporte para Dirección Ejecutiva con avance de proyectos y acciones","Mensual","Coord. Comunicaciones","Interno","Activo","Mensual","X","X","X","X","X","X","X","X","X","X","X","X","Incluir métricas de RRSS y actividades"],
    ["Monitoreo Institucional","Sistematización de métricas RRSS","Consolidado mensual de alcance, engagement e interacciones","Mensual","Coord. Digital","Planilla interna","Activo","Mensual","X","X","X","X","X","X","X","X","X","X","X","X","Compartir con equipo en reunión mensual"],
]

bloques_color = {
    "Redes Sociales Institucionales": "E3F2FD",
    "Boletines y Comunicados":        "F3E5F5",
    "Sitio Web":                      "E8F5E9",
    "Materiales Institucionales":     "FFF8E1",
    "Eventos Institucionales":        "FBE9E7",
    "Monitoreo Institucional":        "F1F8E9",
}

dv_estado6 = DataValidation(type="list",
    formula1='"Activo,En proceso,Planificado,Según evento,Pausado,Completado"',
    showDropDown=False)
ws6.add_data_validation(dv_estado6)

prev_bloque = None
for i, row in enumerate(institucional_data, start=row_inth+1):
    bloque = row[0]
    bg = bloques_color.get(bloque, BLANCO)
    if bloque != prev_bloque:
        # Section divider
        for j in range(1, len(inst_headers)+1):
            c = ws6.cell(row=i, column=j)
            c.fill = fill(VERDE_HOJA) if j == 1 else fill(VERDE_CLARO)
            c.font = font(bold=True, color=BLANCO if j==1 else VERDE_SELVA)
            c.border = border_thin()
        ws6.cell(row=i, column=1).value = bloque
        prev_bloque = bloque
        i_offset = 1
    
    r = i
    for j, val in enumerate(row, start=1):
        c = data_cell(ws6, r, j, val, bg=bg,
                      h="center" if j >= 9 else "left",
                      wrap=(j in [3,21]))
        if j == 7:
            dv_estado6.add(c)
        if j >= 9 and val == "X":
            c.fill = fill(VERDE_LIMA)
            c.font = font(bold=True, color=VERDE_SELVA)
    ws6.row_dimensions[r].height = 20

last_row_inst = row_inth + len(institucional_data)
table_inst = Table(displayName="Inst",
                   ref=f"A{row_inth}:{get_column_letter(len(inst_headers))}{last_row_inst}")
table_inst.tableStyleInfo = TableStyleInfo(name="TableStyleMedium6", showRowStripes=False)
ws6.add_table(table_inst)

ws6.merge_cells(f"A{last_row_inst+2}:U{last_row_inst+4}")
c = ws6.cell(row=last_row_inst+2, column=1,
    value="📌  Para agregar una nueva acción: escribe en la fila siguiente a la tabla con el bloque/categoría correspondiente. "
          "Marca con X los meses de ejecución. El Estado tiene lista desplegable. Puedes filtrar por Bloque para ver solo un tipo de acción.")
c.fill = fill(OCRE_CLARO)
c.font = font(italic=True, color=VERDE_SELVA, size=9)
c.alignment = align("left", "top", wrap=True)
c.border = border_thin()
ws6.row_dimensions[last_row_inst+2].height = 50


# ══════════════════════════════════════════════════════════════════════════════
# TAB COLORS
# ══════════════════════════════════════════════════════════════════════════════
ws1.sheet_properties.tabColor = VERDE_SELVA
ws2.sheet_properties.tabColor = VERDE_HOJA
ws3.sheet_properties.tabColor = OCRE
ws4.sheet_properties.tabColor = AZUL
ws5.sheet_properties.tabColor = TIERRA
ws6.sheet_properties.tabColor = "7B3F9E"

# ── Save ──────────────────────────────────────────────────────────────────────
path = "Matriz_Comunicaciones_AMPA_2026_v2.xlsx"
wb.save(path)
print("Saved:", path)
